"""
Parse company names from Excel and search job boards with per-company queries.
Results are filtered to analyst-style titles and matched to the sheet company name.
"""

import asyncio
import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from app.scrapers.job_scraper import search_jobs
from app.services.location_filter import filter_jobs_by_india_options

# Limit work per request (scraping is slow; many sites rate-limit)
DEFAULT_MAX_COMPANIES = 10
MAX_QUERIES_TOTAL = 24


def _header_row_values(ws) -> List[Any]:
    """First row of sheet as a list (1-based sheet)."""
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


def resolve_company_column_index(header: List[Any]) -> int:
    """0-based column index for company names (default C = 2)."""
    if header:
        for i, cell in enumerate(header):
            if cell is None:
                continue
            if "company" in str(cell).lower().strip():
                return i
    return 2


def resolve_links_column_index(header: List[Any]) -> int:
    """
    0-based column index for role + job links (default F = 5).
    Matches headers like 'Role /Links to apply', 'Links', 'Apply'.
    """
    if not header:
        return 5
    best = None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        s = str(cell).lower()
        if "link" in s or "apply" in s:
            return i
        if "role" in s:
            best = i
    if best is not None:
        return best
    return 5


def _row_key(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).lower().strip())


def extract_companies_from_xlsx(content: bytes) -> List[str]:
    """
    Read .xlsx: use column named 'Company' (any row 1 header), else column C (index 2).
    Skips empty cells and duplicate names (case-insensitive).
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: List[Tuple[Any, ...]] = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return []

    header = list(rows[0]) if rows[0] else []
    company_col = resolve_company_column_index(header)

    data_rows = rows[1:]
    companies: List[str] = []
    seen: Set[str] = set()

    for row in data_rows:
        if company_col >= len(row):
            continue
        val = row[company_col]
        if val is None:
            continue
        name = str(val).strip()
        if len(name) < 2:
            continue
        low = name.lower()
        if low in ("company", "serial", "#", "tier", "name"):
            continue
        if low not in seen:
            seen.add(low)
            companies.append(name)

    return companies


def fill_xlsx_with_job_links(xlsx_bytes: bytes, jobs: List[Dict[str, Any]]) -> bytes:
    """
    Write analyst job titles + URLs into the Role/Links column, one block per company row.
    Preserves the rest of the workbook.
    """
    bio_in = BytesIO(xlsx_bytes)
    wb = load_workbook(bio_in)
    ws = wb.active

    header = _header_row_values(ws)
    company_col = resolve_company_column_index(header) + 1
    links_col = resolve_links_column_index(header) + 1

    by_matched: Dict[str, List[Dict[str, Any]]] = {}
    for j in jobs:
        m = j.get("matched_sheet_company")
        if m:
            k = _row_key(m)
            by_matched.setdefault(k, []).append(j)

    for r in range(2, ws.max_row + 1):
        co = ws.cell(r, company_col).value
        if co is None or str(co).strip() == "":
            continue
        co_str = str(co).strip()
        if co_str.lower() in ("company", "serial", "#", "tier", "name"):
            continue

        key = _row_key(co_str)
        row_jobs: List[Dict[str, Any]] = list(by_matched.get(key, []))
        if not row_jobs:
            row_jobs = [
                j
                for j in jobs
                if company_name_matches(co_str, j.get("company", ""))
                or company_tokens_in_title(co_str, j.get("title", ""))
            ]

        seen_links: Set[str] = set()
        lines: List[str] = []
        for j in row_jobs:
            link = (j.get("link") or "").strip()
            if not link or link == "N/A" or link in seen_links:
                continue
            seen_links.add(link)
            title = (j.get("title") or "Job").replace("\n", " ").strip()
            lines.append(f"{title}\n{link}")

        cell = ws.cell(r, links_col)
        cell.value = "\n\n".join(lines) if lines else ""
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _normalize(s: str) -> str:
    return re.sub(r"\s+", " ", s.lower().strip())


def company_tokens_in_title(sheet_company: str, title: str) -> bool:
    """Fallback when job board lists wrong employer but title mentions the brand."""
    if not sheet_company or not title:
        return False
    tl = title.lower()
    for tok in re.findall(r"[A-Za-z]{4,}", sheet_company):
        if len(tok) >= 4 and tok.lower() in tl:
            return True
    return False


def company_name_matches(sheet_company: str, job_company: str) -> bool:
    """Loose match: substring or strong token overlap (handles HUL vs long legal names)."""
    if not sheet_company or not job_company or job_company == "N/A":
        return False
    a = _normalize(sheet_company)
    b = _normalize(job_company)
    if len(a) < 2 or len(b) < 2:
        return False
    if a in b or b in a:
        return True
    ta = set(re.findall(r"[a-z0-9]{2,}", a))
    tb = set(re.findall(r"[a-z0-9]{2,}", b))
    if not ta or not tb:
        return False
    overlap = ta & tb
    if len(overlap) >= 2:
        return True
    if len(overlap) == 1 and len(ta) <= 3:
        return True
    return False


def title_is_data_or_business_analyst(title: str) -> bool:
    """Keep data / business / BI / reporting analyst roles; drop unrelated 'analyst' noise."""
    if not title or title == "N/A":
        return False
    t = title.lower()
    if re.search(
        r"\b(data\s+analyst|business\s+analyst|bi\s+analyst|analytics\s+analyst|"
        r"reporting\s+analyst|data\s+&\s*business\s+analyst)\b",
        t,
    ):
        return True
    if "analyst" in t and (
        "data" in t or "business" in t or "bi" in t or "analytics" in t or "reporting" in t
    ):
        if "financial" in t and "data" not in t and "business" not in t:
            return False
        return True
    return False


async def search_jobs_for_company_list(
    companies: List[str],
    *,
    max_companies: int = DEFAULT_MAX_COMPANIES,
    concurrency: int = 4,
    india_only: bool = False,
    india_states: Optional[List[str]] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    For each company, run data-analyst and business-analyst queries (bounded concurrency).
    Returns (filtered_jobs, meta).
    """
    take = companies[: max(1, min(max_companies, 50))]
    queries: List[str] = []
    loc_suffix = " India" if india_only else ""
    for c in take:
        queries.append(f"{c} data analyst{loc_suffix}")
        queries.append(f"{c} business analyst{loc_suffix}")

    if len(queries) > MAX_QUERIES_TOTAL:
        queries = queries[:MAX_QUERIES_TOTAL]

    sem = asyncio.Semaphore(concurrency)

    async def one(q: str) -> List[Dict[str, Any]]:
        async with sem:
            try:
                return await search_jobs(q)
            except Exception as e:
                print(f"⚠️ company list query failed {q!r}: {e}")
                return []

    tasks = [one(q) for q in queries]
    chunks = await asyncio.gather(*tasks)

    raw: List[Dict[str, Any]] = []
    seen_links: Set[str] = set()
    for jobs in chunks:
        for job in jobs:
            link = job.get("link") or ""
            if not link or link == "N/A" or link in seen_links:
                continue
            seen_links.add(link)
            raw.append(job)

    matched: List[Dict[str, Any]] = []
    for job in raw:
        title = job.get("title") or ""
        jcompany = job.get("company") or ""
        if not title_is_data_or_business_analyst(title):
            continue
        sheet_match = next((c for c in take if company_name_matches(c, jcompany)), "")
        if not sheet_match:
            sheet_match = next((c for c in take if company_tokens_in_title(c, title)), "")
        if not sheet_match:
            continue
        enriched = dict(job)
        enriched["matched_sheet_company"] = sheet_match
        matched.append(enriched)

    after_analyst = len(matched)
    st_list = [s.strip() for s in (india_states or []) if s and str(s).strip()]
    if india_only or st_list:
        matched = filter_jobs_by_india_options(
            matched,
            india_only=india_only,
            india_states=st_list if st_list else None,
        )

    meta = {
        "companies_requested": len(companies),
        "companies_searched": len(take),
        "queries_run": len(queries),
        "raw_unique": len(raw),
        "matched_after_analyst_filter": after_analyst,
        "matched": len(matched),
        "india_only": india_only,
        "india_states": st_list,
    }
    return matched, meta
